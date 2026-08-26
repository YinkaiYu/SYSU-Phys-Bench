from __future__ import annotations

import csv
import json
from pathlib import Path

from openpyxl import load_workbook

from validate_submissions import FIELDS


ROOT = Path(__file__).resolve().parents[1]
SOURCE_ROOT = ROOT.parent
FILES = {
    "overview": SOURCE_ROOT / "余荫铠-成绩总览.xlsx",
    "details": SOURCE_ROOT / "余荫铠-成绩明细.xlsx",
}
CONTRIBUTOR_ID = "yyk-2020-2024"
TERM_IDS = {
    ("大一", "第一学期"): "2020-fall",
    ("大一", "第二学期"): "2021-spring",
    ("大二", "第一学期"): "2021-fall",
    ("大二", "第二学期"): "2022-spring",
    ("大三", "第一学期"): "2022-fall",
    ("大三", "第二学期"): "2023-spring",
    ("大四", "第一学期"): "2023-fall",
    ("大四", "第二学期"): "2024-spring",
}


def read_rows(path: Path) -> list[list[object]]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    worksheet = workbook.active
    return [
        [cell.value for cell in row]
        for row in worksheet.iter_rows()
        if any(cell.value is not None for cell in row)
    ]


payload = {
    "overview": read_rows(FILES["overview"]),
    "details": read_rows(FILES["details"]),
}

output = "window.GRADE_DATA = " + json.dumps(
    payload,
    ensure_ascii=False,
    separators=(",", ":"),
) + ";\n"
(ROOT / "data.js").write_text(output, encoding="utf-8")


submission_path = ROOT / "data" / "submissions" / f"{CONTRIBUTOR_ID}.csv"
submission_path.parent.mkdir(parents=True, exist_ok=True)
with submission_path.open("w", encoding="utf-8", newline="") as handle:
    writer = csv.DictWriter(handle, fieldnames=FIELDS)
    writer.writeheader()
    for index, row in enumerate(payload["details"][1:], start=1):
        category, course, teacher, academic_year, semester, credits, final_grade, grade_point, rank_value = row
        class_rank, class_size = (int(value) for value in str(rank_value).split("/"))
        writer.writerow({
            "record_id": f"{CONTRIBUTOR_ID}-{index:03d}",
            "contributor_id": CONTRIBUTOR_ID,
            "cohort": 2020,
            "program": "物理学",
            "course_name": course,
            "teacher": teacher or "",
            "category": category,
            "academic_year": academic_year,
            "semester": semester,
            "term_id": TERM_IDS[(academic_year, semester)],
            "credits": credits,
            "final_grade": final_grade,
            "grade_point": grade_point,
            "class_rank": class_rank,
            "class_size": class_size,
        })
